"""XLSX workbook parsing without formula execution or external link loading."""

from __future__ import annotations

import io
import zipfile
from datetime import date, datetime, time
from typing import Any, cast

from akc_cir import BlockType
from defusedxml import ElementTree as SafeElementTree
from defusedxml.common import DefusedXmlException
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.utils.cell import (
    coordinate_to_tuple,
    get_column_letter,
    range_boundaries,
)
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.xml.functions import tostring

from .models import (
    CirBuilder,
    SourceLocation,
    StructuredParseError,
    TableCellSpec,
    normalize_text,
)


def parse_xlsx(data: bytes, builder: CirBuilder) -> str:
    _preflight_worksheet_xml(data, builder)
    try:
        formula_book = load_workbook(
            io.BytesIO(data),
            read_only=False,
            data_only=False,
            keep_links=False,
        )
        value_book = load_workbook(
            io.BytesIO(data),
            read_only=False,
            data_only=True,
            keep_links=False,
        )
    except (InvalidFileException, KeyError, OSError, ValueError, zipfile.BadZipFile) as exc:
        raise StructuredParseError("XLSX_PARSE_FAILED") from exc

    try:
        if len(formula_book.worksheets) > builder.limits.max_sheets:
            raise StructuredParseError("SHEET_LIMIT")
        title = normalize_text(formula_book.properties.title or "")
        sheet_metadata: list[dict[str, Any]] = []
        for sheet_index0, worksheet in enumerate(formula_book.worksheets):
            value_sheet = value_book[worksheet.title]
            hidden = worksheet.sheet_state != "visible"
            sheet_block = builder.add_block(
                block_type=BlockType.HEADING,
                location=SourceLocation(
                    page_index0=sheet_index0,
                    native_object_id=f"xlsx/sheet/{sheet_index0:04d}",
                ),
                raw_text=worksheet.title,
                markdown=f"## {worksheet.title}",
                quality_flags=(("hidden_sheet",) if hidden else ()),
            )
            table_block, formulas = _add_sheet_table(
                worksheet,
                value_sheet,
                builder=builder,
                sheet_index0=sheet_index0,
                parent_id=sheet_block.id,
            )
            explicit_tables = [
                {
                    "name": table.name,
                    "displayName": table.displayName,
                    "ref": table.ref,
                }
                for table in sorted(
                    worksheet.tables.values(),
                    key=lambda item: (item.ref, item.name),
                )
            ]
            image_count, chart_count = _add_sheet_assets(
                worksheet,
                builder=builder,
                sheet_index0=sheet_index0,
                parent_id=sheet_block.id,
            )
            hidden_rows = sorted(
                int(index)
                for index, dimension in worksheet.row_dimensions.items()
                if dimension.hidden
            )
            hidden_columns = sorted(
                str(index)
                for index, dimension in worksheet.column_dimensions.items()
                if dimension.hidden
            )
            sheet_metadata.append(
                {
                    "pageIndex0": sheet_index0,
                    "name": worksheet.title,
                    "state": worksheet.sheet_state,
                    "tables": explicit_tables,
                    "imageCount": image_count,
                    "chartCount": chart_count,
                    "formulas": formulas,
                    "hiddenRows": hidden_rows,
                    "hiddenColumns": hidden_columns,
                    "hasCanonicalTable": table_block is not None,
                }
            )
        builder.metadata["sheets"] = sheet_metadata
        if not builder.blocks:
            raise StructuredParseError("XLSX_EMPTY_WORKBOOK")
        return title or _filename_title(builder.source_filename)
    finally:
        formula_book.close()
        value_book.close()


def _add_sheet_table(
    worksheet: Any,
    value_sheet: Any,
    *,
    builder: CirBuilder,
    sheet_index0: int,
    parent_id: str,
) -> tuple[str | None, list[dict[str, Any]]]:
    actual_cells: dict[tuple[int, int], Cell] = {
        coordinate: cell
        for coordinate, cell in worksheet._cells.items()
        if isinstance(cell, Cell) and cell.value is not None
    }
    merge_anchors: dict[tuple[int, int], tuple[int, int]] = {}
    covered: set[tuple[int, int]] = set()
    for merged_range in worksheet.merged_cells.ranges:
        min_column, min_row, max_column, max_row = _range_boundaries(str(merged_range))
        merge_anchors[(min_row, min_column)] = (
            max_row - min_row + 1,
            max_column - min_column + 1,
        )
        for row in range(min_row, max_row + 1):
            for column in range(min_column, max_column + 1):
                if (row, column) != (min_row, min_column):
                    covered.add((row, column))

    coordinates = (set(actual_cells) | set(merge_anchors)) - covered
    if not coordinates:
        return None, []
    if len(coordinates) > builder.limits.max_cells_per_sheet:
        raise StructuredParseError("SHEET_CELL_LIMIT")
    min_row = min(row for row, _column in coordinates)
    max_row = max(row for row, _column in coordinates)
    min_column = min(column for _row, column in coordinates)
    max_column = max(column for _row, column in coordinates)
    row_count = max_row - min_row + 1
    column_count = max_column - min_column + 1
    if row_count > builder.limits.max_rows_per_sheet:
        raise StructuredParseError("SHEET_ROW_LIMIT")
    if column_count > builder.limits.max_columns_per_sheet:
        raise StructuredParseError("SHEET_COLUMN_LIMIT")

    specs: list[TableCellSpec] = []
    formula_count = 0
    formulas: list[dict[str, Any]] = []
    for row, column in sorted(coordinates):
        cell = actual_cells.get((row, column))
        raw_value = cell.value if cell is not None else None
        raw_text = _cell_text(raw_value)
        normalized_text = raw_text
        quality_flags: list[str] = []
        if cell is not None and cell.data_type == "f":
            formula_count += 1
            cached_value = value_sheet.cell(row=row, column=column).value
            normalized_text = _cell_text(cached_value) if cached_value is not None else raw_text
            quality_flags.append("formula_preserved_not_executed")
            if cached_value is None:
                quality_flags.append("formula_cached_value_missing")
        if worksheet.row_dimensions[row].hidden:
            quality_flags.append("hidden_row")
        column_letter = get_column_letter(column)
        if worksheet.column_dimensions[column_letter].hidden:
            quality_flags.append("hidden_column")
        row_span, column_span = merge_anchors.get((row, column), (1, 1))
        coordinate = f"{column_letter}{row}"
        if cell is not None and cell.data_type == "f":
            formulas.append(
                {
                    "cell": coordinate,
                    "formula": raw_text,
                    "cachedValue": (
                        _cell_text(value_sheet.cell(row=row, column=column).value)
                        if value_sheet.cell(row=row, column=column).value is not None
                        else None
                    ),
                    "cachedValuePresent": (
                        value_sheet.cell(row=row, column=column).value is not None
                    ),
                }
            )
        specs.append(
            TableCellSpec(
                row_index0=row - min_row,
                column_index0=column - min_column,
                row_span=row_span,
                column_span=column_span,
                raw_text=raw_text,
                normalized_text=normalized_text,
                location=SourceLocation(
                    page_index0=sheet_index0,
                    native_object_id=(f"xlsx/sheet/{sheet_index0:04d}/cell/{coordinate}"),
                ),
                quality_flags=tuple(quality_flags),
            )
        )

    if formula_count:
        builder.add_warning("xlsx_formulas_not_executed")
    range_ref = f"{get_column_letter(min_column)}{min_row}:{get_column_letter(max_column)}{max_row}"
    explicit_header = any(
        _range_boundaries(table.ref)[1] == min_row for table in worksheet.tables.values()
    )
    first_row_specs = [cell for cell in specs if cell.row_index0 == 0]
    inferred_header = bool(first_row_specs) and all(
        cell.raw_text.strip() for cell in first_row_specs
    )
    header_row_count = 1 if explicit_header or inferred_header else 0
    table_block = builder.add_table(
        location=SourceLocation(
            page_index0=sheet_index0,
            native_object_id=(f"xlsx/sheet/{sheet_index0:04d}/range/{range_ref}"),
        ),
        row_count=row_count,
        column_count=column_count,
        cells=tuple(specs),
        header_row_count=header_row_count,
        parent_id=parent_id,
        quality_flags=(("header_row_inferred",) if inferred_header and not explicit_header else ()),
    )
    return table_block.id, formulas


def _add_sheet_assets(
    worksheet: Any,
    *,
    builder: CirBuilder,
    sheet_index0: int,
    parent_id: str,
) -> tuple[int, int]:
    images = tuple(getattr(worksheet, "_images", ()))
    charts = tuple(getattr(worksheet, "_charts", ()))
    for image_index, image in enumerate(images):
        payload = bytes(image._data())
        native_part = str(getattr(image, "path", f"/xl/media/image{image_index + 1}")).lstrip("/")
        anchor = _anchor_descriptor(getattr(image, "anchor", None))
        asset_id = builder.register_asset(
            payload=payload,
            native_part=native_part,
            media_type=_image_media_type(str(getattr(image, "format", ""))),
            kind="image",
            filename=native_part.rsplit("/", 1)[-1],
            width_px=round(float(getattr(image, "width", 0))) or None,
            height_px=round(float(getattr(image, "height", 0))) or None,
            metadata={
                "sheet": worksheet.title,
                "anchor": anchor,
            },
        )
        label = normalize_text(str(getattr(image, "_name", native_part.rsplit("/", 1)[-1])))
        builder.add_block(
            block_type=BlockType.FIGURE,
            location=SourceLocation(
                page_index0=sheet_index0,
                native_object_id=(f"xlsx/sheet/{sheet_index0:04d}/image/{image_index:04d}"),
                image_asset_id=asset_id,
            ),
            raw_text=label,
            markdown=f"![{label}](akc-asset:{asset_id})",
            parent_id=parent_id,
            quality_flags=("embedded_asset_extracted",),
        )

    for chart_index, chart in enumerate(charts):
        chart_xml = bytes(tostring(chart._write()))
        native_part = str(getattr(chart, "path", f"/xl/charts/chart{chart_index + 1}.xml")).lstrip(
            "/"
        )
        anchor = _anchor_descriptor(getattr(chart, "anchor", None))
        chart_data = _xlsx_chart_data(chart)
        asset_id = builder.register_asset(
            payload=chart_xml,
            native_part=native_part,
            media_type=("application/vnd.openxmlformats-officedocument.drawingml.chart+xml"),
            kind="chart",
            filename=native_part.rsplit("/", 1)[-1],
            metadata={
                "sheet": worksheet.title,
                "anchor": anchor,
                "chartData": chart_data,
            },
        )
        label = normalize_text(str(chart_data["title"])) or f"Chart {chart_index + 1}"
        series_lines = [
            f"{series['name']}: {series['valueReference']}"
            for series in chart_data["series"]
            if series.get("valueReference")
        ]
        builder.add_block(
            block_type=BlockType.FIGURE,
            location=SourceLocation(
                page_index0=sheet_index0,
                native_object_id=(f"xlsx/sheet/{sheet_index0:04d}/chart/{chart_index:04d}"),
                image_asset_id=asset_id,
            ),
            raw_text="\n".join((label, *series_lines)),
            markdown=f"**Chart: {label}**\n\n" + "\n".join(series_lines),
            parent_id=parent_id,
            quality_flags=(
                "chart_structure_extracted",
                "embedded_asset_extracted",
            ),
        )
    return len(images), len(charts)


def _anchor_descriptor(anchor: Any) -> dict[str, int] | str | None:
    if anchor is None:
        return None
    if isinstance(anchor, str):
        return anchor
    start = getattr(anchor, "_from", None)
    end = getattr(anchor, "to", None)
    if start is None:
        return normalize_text(str(anchor))
    descriptor = {
        "fromRow0": int(start.row),
        "fromColumn0": int(start.col),
        "fromRowOffsetEmu": int(start.rowOff),
        "fromColumnOffsetEmu": int(start.colOff),
    }
    if end is not None:
        descriptor.update(
            {
                "toRow0": int(end.row),
                "toColumn0": int(end.col),
                "toRowOffsetEmu": int(end.rowOff),
                "toColumnOffsetEmu": int(end.colOff),
            }
        )
    return descriptor


def _image_media_type(image_format: str) -> str:
    normalized = image_format.casefold().lstrip(".")
    return {
        "jpg": "image/jpeg",
        "jpeg": "image/jpeg",
        "png": "image/png",
        "gif": "image/gif",
        "bmp": "image/bmp",
        "tif": "image/tiff",
        "tiff": "image/tiff",
        "webp": "image/webp",
    }.get(normalized, "application/octet-stream")


def _xlsx_chart_data(chart: Any) -> dict[str, Any]:
    series_data: list[dict[str, Any]] = []
    for series_index, series in enumerate(chart.series):
        title_reference = _reference_formula(getattr(series, "tx", None))
        value_reference = _reference_formula(
            getattr(series, "val", None) or getattr(series, "yVal", None)
        )
        category_reference = _reference_formula(
            getattr(series, "cat", None) or getattr(series, "xVal", None)
        )
        series_data.append(
            {
                "seriesIndex0": series_index,
                "name": title_reference or f"Series {series_index + 1}",
                "titleReference": title_reference,
                "categoryReference": category_reference,
                "valueReference": value_reference,
            }
        )
    return {
        "title": _xlsx_chart_title(chart),
        "chartType": type(chart).__name__,
        "series": series_data,
    }


def _reference_formula(reference: Any) -> str | None:
    if reference is None:
        return None
    for candidate_name in ("strRef", "numRef"):
        candidate = getattr(reference, candidate_name, None)
        formula = getattr(candidate, "f", None)
        if formula:
            return normalize_text(str(formula))
    formula = getattr(reference, "f", None)
    return normalize_text(str(formula)) if formula else None


def _xlsx_chart_title(chart: Any) -> str:
    title = getattr(chart, "title", None)
    if title is None:
        return ""
    text: list[str] = []
    try:
        paragraphs = title.tx.rich.p
    except AttributeError:
        return ""
    for paragraph in paragraphs:
        for run in getattr(paragraph, "r", ()):
            value = getattr(run, "t", None)
            if value:
                text.append(str(value))
        end_value = getattr(paragraph, "endParaRPr", None)
        if end_value is not None and getattr(paragraph, "t", None):
            text.append(str(paragraph.t))
    return normalize_text("".join(text))


def _preflight_worksheet_xml(data: bytes, builder: CirBuilder) -> None:
    try:
        archive = zipfile.ZipFile(io.BytesIO(data))
    except (OSError, zipfile.BadZipFile) as exc:
        raise StructuredParseError("XLSX_PARSE_FAILED") from exc
    with archive:
        worksheet_entries = sorted(
            (
                entry
                for entry in archive.infolist()
                if entry.filename.replace("\\", "/").casefold().startswith("xl/worksheets/")
                and entry.filename.casefold().endswith(".xml")
            ),
            key=lambda entry: entry.filename,
        )
        if len(worksheet_entries) > builder.limits.max_sheets:
            raise StructuredParseError("SHEET_LIMIT")
        for entry in worksheet_entries:
            try:
                root = SafeElementTree.fromstring(archive.read(entry))
            except (DefusedXmlException, SafeElementTree.ParseError) as exc:
                raise StructuredParseError("XLSX_UNSAFE_XML") from exc
            cell_count = 0
            merged_area = 0
            rows: set[int] = set()
            for element in root.iter():
                local_name = element.tag.rsplit("}", 1)[-1]
                if local_name == "c":
                    cell_count += 1
                    reference = element.attrib.get("r")
                    if reference:
                        row, column = _safe_coordinate(reference)
                        rows.add(row)
                        _enforce_sheet_bounds(row, column, builder)
                elif local_name == "row":
                    raw_row = element.attrib.get("r")
                    if raw_row and raw_row.isdigit():
                        rows.add(int(raw_row))
                elif local_name == "dimension":
                    reference = element.attrib.get("ref")
                    if reference:
                        _enforce_range(reference, builder)
                elif local_name == "mergeCell":
                    reference = element.attrib.get("ref")
                    if reference:
                        merged_area += _enforce_range(reference, builder)
                        if merged_area > builder.limits.max_cells_per_sheet:
                            raise StructuredParseError("SHEET_CELL_LIMIT")
            if cell_count > builder.limits.max_cells_per_sheet:
                raise StructuredParseError("SHEET_CELL_LIMIT")
            if len(rows) > builder.limits.max_rows_per_sheet:
                raise StructuredParseError("SHEET_ROW_LIMIT")


def _safe_coordinate(reference: str) -> tuple[int, int]:
    try:
        row, column = coordinate_to_tuple(reference.replace("$", ""))
    except (TypeError, ValueError) as exc:
        raise StructuredParseError("XLSX_INVALID_CELL_REFERENCE") from exc
    return row, column


def _enforce_range(reference: str, builder: CirBuilder) -> int:
    try:
        min_column, min_row, max_column, max_row = _range_boundaries(reference.replace("$", ""))
    except (TypeError, ValueError) as exc:
        raise StructuredParseError("XLSX_INVALID_RANGE") from exc
    _enforce_sheet_bounds(max_row, max_column, builder)
    if max_row - min_row + 1 > builder.limits.max_rows_per_sheet:
        raise StructuredParseError("SHEET_ROW_LIMIT")
    if max_column - min_column + 1 > builder.limits.max_columns_per_sheet:
        raise StructuredParseError("SHEET_COLUMN_LIMIT")
    area = (max_row - min_row + 1) * (max_column - min_column + 1)
    if area > builder.limits.max_cells_per_sheet:
        raise StructuredParseError("SHEET_CELL_LIMIT")
    return area


def _enforce_sheet_bounds(row: int, column: int, builder: CirBuilder) -> None:
    if row > builder.limits.max_rows_per_sheet:
        raise StructuredParseError("SHEET_ROW_LIMIT")
    if column > builder.limits.max_columns_per_sheet:
        raise StructuredParseError("SHEET_COLUMN_LIMIT")


def _range_boundaries(reference: str) -> tuple[int, int, int, int]:
    try:
        boundaries = range_boundaries(reference)
    except (TypeError, ValueError) as exc:
        raise StructuredParseError("XLSX_INVALID_RANGE") from exc
    if any(value is None for value in boundaries):
        raise StructuredParseError("XLSX_INVALID_RANGE")
    return cast(tuple[int, int, int, int], boundaries)


def _cell_text(value: Any) -> str:
    if value is None or isinstance(value, MergedCell):
        return ""
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    return normalize_text(str(value))


def _filename_title(filename: str) -> str:
    return normalize_text(filename.rsplit(".", 1)[0].replace("_", " ")) or "Workbook"
